#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import print_function #,unicode_literals

USE_XLRD=0

#this puts ints as floats :/ and formatting_info is not there yet
def letter2index0_openpyxl( a):
    import xlrd.xlsx
    r = xlrd.xlsx._UPPERCASE_1_REL_INDEX[a]
    assert r>0
    return r-1

#if 123 is formatted under 0000 -> manual reformat-to-text
def letter2index0_xlrd( a):
    import openpyxl.utils
    r = openpyxl.utils.column_index_from_string( a)
    assert r>0
    return r-1

ALLSHEETS = object()
def read( fname, sheets =[], as_sheet =False):
    use_xlrd = USE_XLRD
    if not use_xlrd:
      try:
        import openpyxl
        wb = openpyxl.load_workbook( filename= fname,
                    read_only= True,
                    #guess_types= False,
                    data_only= True,    #no formulas
                    )
        wsnames = wb.get_sheet_names()
        letter2index0 = letter2index0_openpyxl
      except openpyxl.utils.exceptions.InvalidFileException:
        use_xlrd = 1
    if use_xlrd:
        import xlrd
        wb = xlrd.open_workbook( filename= fname, formatting_info= True )   # no formatting_info yet :(
        wsnames = wb.sheet_names()
        letter2index0 = letter2index0_xlrd
    if not sheets:
        if len(wsnames) >1:
            print( 'cannot choose, more than one available sheets:')
            for ws in wsnames: print( ' ', repr(ws))
            raise SystemExit(1)

        wsnames = [ wsnames[0] ]
    elif sheets is not ALLSHEETS:
        wsnames = sheets

    for sname in wsnames:
        if use_xlrd:
            ws = wb.sheet_by_name( sname)
            rows = ws.get_rows()

            def value( cell):
                if isinstance( cell.value, float) and cell.value.is_integer():
                    return int( cell.value)
                return cell.value
        else:
            ws = wb[ sname ]
            rows = ws.rows

            def value( cell):
                if cell.value is not None:
                    if cell.number_format and cell.number_format.startswith('0'):
                        return str(cell.value).zfill( len( cell.number_format ))
                return cell.value

        if not as_sheet:
            ws = [ [ value( cell) for cell in row]
                    for row in rows ]
        yield sname.strip(), ws


def strip( array, where =[]): #top =False, left =False, right =True, bottom =True):
    minx= miny= 0
    maxw= maxh = -1
    if 'top' in where:
        i=0
        for i,row in enumerate( array):
            if any( cell is not None for cell in row): break
        miny = i
        array[ :i] = []

    if 'left' in where:
        minx = 0
        for row in array:
            i=0
            for i,cell in enumerate( row):
                if cell is not None: break
            minx = min( minx, i)
        for row in array:
            row[ :minx] = []

    if 'right' in where:
        maxw = 0
        for row in array:
            i=0
            for i,cell in enumerate( row[::-1]):
                if cell is not None: break
            w = len(row)-i
            maxw = max( maxw, w)
        for row in array:
            row[ maxw:] = []

    if 'bottom' in where:
        i=0
        for i,row in enumerate( array[::-1]):
            if any( cell is not None for cell in row): break
        maxh = len(array)-i
        array[ maxh:] = []

    print( 'x', minx, 'y', miny, 'w', maxw, 'h', maxh)
    return array

def select_columns( array, columns =()):
    if not columns: return
    onlythese = columns
    #needs_names = not columns[0].isdigit()
    #if not needs_names:
    onlythese = [ int(i) if i.isdigit() else letter2index0(i) for i in onlythese ]
    print( onlythese )

    for row in array:
        #if needs_names:
        #    needs_names = False
        #    onlythese = [ i for i,c in enumerate( row) if c in onlythese ]
        if onlythese: row[:] = [ row[i] for i in onlythese ]
    return array

import csv
def write( fname, array, delimiter =None):
    #with open( fname, 'wb') as f: nonono
    try:
        f = open( fname, 'x', newline='')
    except FileExistsError:
        print( '! file exists', fname)
        return
    print( '>', fname)
    with f: #open( fname, 'w', newline='') as f:
        c = csv.writer(f, **(delimiter and dict( delimiter=delimiter) or {}))
        for row in array:
            c.writerow( row )

if __name__ == '__main__':
    from svd_util import optz, osextra
    optz.help( '''
extract excel sheet(s) into .csv
args-in-any-order: file.xlsx [sheetname] [sheetname] ..
 unless --allsheets, no sheet-names means first if only one, else shows all sheet-names and stops
 examples:
    ... path/to/somename.xls --all --dir --path .   -->> dumps all sheets as ./somename/sheetname*.csv

'''.strip())
    optz.str( 'csv',        help= 'name of output file.csv ; used as prefix if allsheets; defaults to path+name of file.xlsx + .sheetname-if-specified')
    optz.str( 'path',       help= 'output path, default is whereever .xls is')
    optz.bool( 'allsheets', help= 'save all sheets, named csvname.sheetname.csv')
    optz.bool( 'dir_from_name',  help= 'create outname as folder, i.e. csvname/sheetname.csv')
#    optz.bool( 'nostrip',   help= 'do not strip empty columns/rows at right/bottom - see also --strip_top_left')
    optz.str( 'strip',      help= 'strip empty columns/rows at where - left,top,bottom,right, comma separated; default [%default]', default= 'right,bottom')
    optz.str( 'columns',    help= 'columns to extract, comma separated (index-0-based, or excel letter A-...) ; default to all')  #name if first row of result .csv contain names) ; may be multiple
    optz.str( 'delimiter',  help= '.csv column delimiter, default is comma ","')
    oparser = optz.oparser
    optz,argz = optz.get()
    fcsv = optz.csv
    sheets = []
    fxls = None
    for a in argz:
        if a.lower().endswith( '.xls') or a.lower().endswith( '.xlsx'):
            fxls = a
        else: sheets.append( a )

    if not fxls:
        oparser.error( 'needs .xls or .xlsx file')
    if optz.allsheets:
        if sheets: oparser.error( 'needs either --allsheets or sheetnames, not both')
        if fcsv:  print( 'auto-appending sheetnames to .csv name')
        sheets = ALLSHEETS

    from os.path import join,basename
    if not fcsv:
        fcsv = osextra.withoutext( fxls, '.xlsx', '.xls', ignorecase=True)
    else:
        fcsv = osextra.withoutext( fcsv, '.csv', ignorecase=True)
    if optz.path:
        osextra.makedirs( optz.path)
        fcsv = join( optz.path, basename( fcsv))
    if optz.dir_from_name:
        osextra.makedirs( fcsv)
    for sheetname, data in read( fxls, sheets):
        if optz.strip:      strip( data, optz.strip.split(',') )
        if optz.columns:    select_columns( data, optz.columns.split(','))
        afcsv = fcsv
        if sheets:
            afcsv += ('/' if optz.dir_from_name else '.') + sheetname
        write( afcsv + '.csv', data, delimiter= optz.delimiter)

# vim:ts=4:sw=4:expandtab
